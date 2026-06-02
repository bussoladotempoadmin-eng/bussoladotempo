"""
Gera 13-planilha-piloto.xlsx com fórmulas vivas (SUMIFS, IF).
Editar qualquer célula recalcula matriz e insights automaticamente no Excel/Google Sheets.
Framework: BÚSSOLA DO TEMPO — categorias: Importante / Urgente / Disperso.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.comments import Comment

# Caminho de saída
OUT = r"c:\Users\Doctum\projetos\Agenda Lucas Silveira\13-planilha-piloto.xlsx"

# Cores (hex sem #)
DOCTUM = "3B82F6"
TRIBO = "F97316"
BRUNA = "A855F7"
CUIDAJA = "22C55E"
IMP = "10B981"
URG = "EF4444"
DISP = "94A3B8"
HEADER_BG = "1E293B"
HEADER_FG = "FFFFFF"
ACCENT = "3B82F6"
SOFT_BG = "F1F5F9"

thin = Side(border_style="thin", color="E2E8F0")
border_all = Border(top=thin, bottom=thin, left=thin, right=thin)

def header_fill(color=HEADER_BG):
    return PatternFill(fill_type="solid", start_color=color, end_color=color)

def soft_fill(color=SOFT_BG):
    return PatternFill(fill_type="solid", start_color=color, end_color=color)

def header_font(color=HEADER_FG, bold=True, size=11):
    return Font(color=color, bold=bold, size=size, name="Calibri")

# Dados pré-preenchidos (mesma lógica do HTML)
FRENTES = [
    ("Doctum", "🏢", DOCTUM, 36),
    ("Tribo", "🛒", TRIBO, 18),
    ("Dra. Bruna", "🎓", BRUNA, 2.5),
    ("CuidaJA", "🤝", CUIDAJA, 2.5),
]

BLOCOS = [
    ("SEG", "08:00", "12:00", "Doctum", "Revisar pipeline + alinhar marketing", "Importante", "Importante"),
    ("SEG", "13:30", "18:00", "Doctum", "1:1 vendedor X + ajustes", "Urgente", "Urgente"),
    ("SEG", "18:30", "21:30", "Tribo",  "🎤 Mentoria ao vivo - aula 04", "Importante", "Importante"),
    ("TER", "08:00", "12:00", "Doctum", "Reunião comitê + revisar metas", "Importante", "Urgente"),
    ("TER", "13:30", "14:00", "Dra. Bruna", "Status semanal Dra. Bruna", "Disperso", "Disperso"),
    ("TER", "14:00", "14:30", "CuidaJA", "Status CuidaJA", "Disperso", "Disperso"),
    ("TER", "14:30", "17:00", "Tribo",  "Gravar aula 05 mentoria", "Importante", "Importante"),
    ("TER", "17:00", "19:00", "Doctum", "Email + alinhamentos", "Disperso", "Disperso"),
    ("QUA", "08:00", "12:00", "Doctum", "Reuniões com clientes + forecast", "Importante", "Urgente"),
    ("QUA", "13:30", "14:00", "Dra. Bruna", "Revisar material aula Dra. Bruna", "Importante", "Importante"),
    ("QUA", "14:00", "14:30", "CuidaJA", "CuidaJA - mensagens / alinhamento", "Disperso", "Disperso"),
    ("QUA", "14:30", "17:00", "Tribo",  "TriboCRM - revisar funcionalidades", "Importante", "Importante"),
    ("QUA", "17:00", "19:00", "Doctum", "Doctum - administrativo", "Disperso", "Disperso"),
    ("QUI", "08:00", "12:00", "Doctum", "Workshop com time de vendas", "Importante", "Importante"),
    ("QUI", "13:30", "17:30", "Doctum", "Reuniões 1:1 + planejamento", "Importante", "Urgente"),
    ("QUI", "18:00", "21:00", "Tribo",  "📺 Live Tribo", "Importante", "Importante"),
    ("SEX", "08:00", "12:00", "Doctum", "Fechamento semana + relatórios", "Importante", "Urgente"),
    ("SEX", "13:30", "14:30", "Dra. Bruna", "Reunião Dra. Bruna - revisão semana", "Importante", "Importante"),
    ("SEX", "14:30", "15:30", "CuidaJA", "Reunião CuidaJA - revisão semana", "Importante", "Importante"),
    ("SEX", "15:30", "19:00", "Doctum", "Doctum - fechamento + emails", "Urgente", "Urgente"),
    ("SÁB", "08:00", "13:00", "Tribo",  "Tribo - estratégia + produto", "Importante", "Importante"),
    ("SÁB", "13:00", "13:30", "Dra. Bruna", "Bruna - preparar semana", "Importante", "Importante"),
    ("SÁB", "13:30", "14:00", "CuidaJA", "CuidaJA - preparar semana", "Importante", "Importante"),
    ("SÁB", "14:00", "16:00", "Tribo",  "Tribo - conteúdo / produto", "Importante", "Importante"),
]

DIAS_VALIDOS = "SEG,TER,QUA,QUI,SEX,SÁB,DOM"
TRIADE_OPTS = "Importante,Urgente,Disperso"

wb = Workbook()

# ============================================
# ABA 1: SOBRE
# ============================================
ws = wb.active
ws.title = "Sobre"
ws.column_dimensions["A"].width = 90

linhas = [
    ("🧭 Planilha Piloto · App de Gestão de Tempo", 18, True, HEADER_BG),
    ("Matriz Frente × Bússola do Tempo", 12, False, None),
    ("", 11, False, None),
    ("Como usar", 14, True, ACCENT),
    ("1. Vá pra aba 'Frentes' e ajuste seu orçamento de horas por frente.", 11, False, None),
    ("2. Vá pra aba 'Blocos' e edite/adicione os blocos da sua semana.", 11, False, None),
    ("3. Na coluna 'Planejado', classifique como Importante/Urgente/Disperso antes da semana.", 11, False, None),
    ("4. Na coluna 'Realizado', ajuste depois (durante ou no fim da semana) conforme o que aconteceu de verdade.", 11, False, None),
    ("5. Veja a aba 'Espelho' — matriz calculada automaticamente.", 11, False, None),
    ("6. Veja a aba 'Insights' — padrões detectados automaticamente.", 11, False, None),
    ("", 11, False, None),
    ("Importante / Urgente / Disperso — a Bússola do Tempo", 14, True, ACCENT),
    ("🎯 Importante: estratégico, alinhado com objetivos, gera resultado real.", 11, False, None),
    ("🔥 Urgente: tem prazo apertado, exige ação imediata — pode não ser importante.", 11, False, None),
    ("💨 Disperso: parece trabalho mas não gera resultado nem tem prazo real.", 11, False, None),
    ("", 11, False, None),
    ("Atalhos úteis", 14, True, ACCENT),
    ("• Tudo é calculado por fórmula. Mude qualquer célula e veja a matriz/insights recalcularem.", 11, False, None),
    ("• Pra adicionar bloco: vá pra última linha da aba 'Blocos' e digite. Fórmulas se estendem sozinhas.", 11, False, None),
    ("• Pra adicionar frente: vá pra última linha da aba 'Frentes' e digite. Atualize a aba 'Espelho' (instruções lá).", 11, False, None),
    ("• Drop-down nas colunas Frente / Planejado / Realizado.", 11, False, None),
    ("", 11, False, None),
    ("Versão V1 · Gerada 25/05/2026 · Caso Lucas Silveira pré-preenchido", 10, False, None),
]

for i, (txt, size, bold, fill) in enumerate(linhas, start=1):
    cell = ws.cell(row=i, column=1, value=txt)
    cell.font = Font(size=size, bold=bold, color=(HEADER_FG if fill else "0F172A"))
    if fill:
        cell.fill = header_fill(fill)
    cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[i].height = max(20, size + 8)

# ============================================
# ABA 2: FRENTES
# ============================================
ws = wb.create_sheet("Frentes")
headers = ["Nome", "Ícone", "Orçamento (h/sem)", "Realizado (h)", "Variação", "Status"]
for col, h in enumerate(headers, start=1):
    c = ws.cell(row=1, column=col, value=h)
    c.font = header_font()
    c.fill = header_fill()
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = border_all

for i, (nome, icon, cor, orc) in enumerate(FRENTES, start=2):
    ws.cell(row=i, column=1, value=nome).font = Font(bold=True, size=12)
    ws.cell(row=i, column=2, value=icon).alignment = Alignment(horizontal="center")
    ws.cell(row=i, column=3, value=orc).alignment = Alignment(horizontal="center")
    # Realizado: SUMIF na coluna D (Duração) da aba Blocos onde E (Frente) = este nome
    ws.cell(row=i, column=4, value=f'=SUMIF(Blocos!E:E,A{i},Blocos!D:D)').alignment = Alignment(horizontal="center")
    # Variação = Realizado - Orçamento
    ws.cell(row=i, column=5, value=f'=D{i}-C{i}').alignment = Alignment(horizontal="center")
    # Status: texto baseado em variação
    ws.cell(row=i, column=6, value=f'=IF(ABS(E{i})<1,"✓ no plano",IF(E{i}>0,"⬆ estourou","⬇ abaixo"))').alignment = Alignment(horizontal="center")

    # Aplicar cor da frente na primeira célula
    ws.cell(row=i, column=1).fill = PatternFill(fill_type="solid", start_color=cor, end_color=cor)
    ws.cell(row=i, column=1).font = Font(bold=True, size=12, color=HEADER_FG)

    for col in range(1, 7):
        ws.cell(row=i, column=col).border = border_all

# Largura das colunas
widths = {"A": 22, "B": 10, "C": 22, "D": 18, "E": 14, "F": 18}
for col, w in widths.items():
    ws.column_dimensions[col].width = w

ws.row_dimensions[1].height = 28

# Adicionar dica
ws.cell(row=len(FRENTES) + 4, column=1, value="💡 Pra adicionar frente: digite o nome em A{}, ícone em B, orçamento em C. As fórmulas se replicam.".format(len(FRENTES) + 4)).font = Font(italic=True, size=10, color="64748B")

# ============================================
# ABA 3: BLOCOS
# ============================================
ws = wb.create_sheet("Blocos")
headers = ["Dia", "Início", "Fim", "Duração (h)", "Frente", "Tarefa / entrega", "Planejado", "Realizado"]
for col, h in enumerate(headers, start=1):
    c = ws.cell(row=1, column=col, value=h)
    c.font = header_font()
    c.fill = header_fill()
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = border_all

for i, (dia, ini, fim, frente, tarefa, plan, real) in enumerate(BLOCOS, start=2):
    ws.cell(row=i, column=1, value=dia).alignment = Alignment(horizontal="center")
    ws.cell(row=i, column=2, value=ini).alignment = Alignment(horizontal="center")
    ws.cell(row=i, column=3, value=fim).alignment = Alignment(horizontal="center")
    # Duração calculada
    ws.cell(row=i, column=4, value=f'=IFERROR((TIMEVALUE(C{i})-TIMEVALUE(B{i}))*24,"")').alignment = Alignment(horizontal="center")
    ws.cell(row=i, column=5, value=frente)
    ws.cell(row=i, column=6, value=tarefa)
    ws.cell(row=i, column=7, value=plan).alignment = Alignment(horizontal="center")
    ws.cell(row=i, column=8, value=real).alignment = Alignment(horizontal="center")

    for col in range(1, 9):
        ws.cell(row=i, column=col).border = border_all

# Larguras
widths = {"A": 8, "B": 10, "C": 10, "D": 12, "E": 14, "F": 42, "G": 16, "H": 16}
for col, w in widths.items():
    ws.column_dimensions[col].width = w

ws.row_dimensions[1].height = 28

# Validações de dados
N = 200  # linhas pra cobrir crescimento
dv_dia = DataValidation(type="list", formula1=f'"{DIAS_VALIDOS}"', allow_blank=True)
dv_dia.add(f"A2:A{N}")
ws.add_data_validation(dv_dia)

dv_frente = DataValidation(type="list", formula1="=Frentes!$A$2:$A$50", allow_blank=True)
dv_frente.add(f"E2:E{N}")
ws.add_data_validation(dv_frente)

dv_triade = DataValidation(type="list", formula1=f'"{TRIADE_OPTS}"', allow_blank=True)
dv_triade.add(f"G2:G{N}")
dv_triade.add(f"H2:H{N}")
ws.add_data_validation(dv_triade)

# Formatação condicional pra colorir Bússola
def add_text_rule(rng, text, color):
    rule = FormulaRule(formula=[f'EXACT(G2,"{text}")'], fill=PatternFill(fill_type="solid", start_color=color, end_color=color), font=Font(color="FFFFFF", bold=True))
    return rule

# Importante = verde
ws.conditional_formatting.add(f"G2:G{N}", FormulaRule(formula=[f'EXACT($G2,"Importante")'], fill=PatternFill(fill_type="solid", start_color=IMP, end_color=IMP), font=Font(color="FFFFFF", bold=True)))
ws.conditional_formatting.add(f"G2:G{N}", FormulaRule(formula=[f'EXACT($G2,"Urgente")'], fill=PatternFill(fill_type="solid", start_color=URG, end_color=URG), font=Font(color="FFFFFF", bold=True)))
ws.conditional_formatting.add(f"G2:G{N}", FormulaRule(formula=[f'EXACT($G2,"Disperso")'], fill=PatternFill(fill_type="solid", start_color=DISP, end_color=DISP), font=Font(color="FFFFFF", bold=True)))

ws.conditional_formatting.add(f"H2:H{N}", FormulaRule(formula=[f'EXACT($H2,"Importante")'], fill=PatternFill(fill_type="solid", start_color=IMP, end_color=IMP), font=Font(color="FFFFFF", bold=True)))
ws.conditional_formatting.add(f"H2:H{N}", FormulaRule(formula=[f'EXACT($H2,"Urgente")'], fill=PatternFill(fill_type="solid", start_color=URG, end_color=URG), font=Font(color="FFFFFF", bold=True)))
ws.conditional_formatting.add(f"H2:H{N}", FormulaRule(formula=[f'EXACT($H2,"Disperso")'], fill=PatternFill(fill_type="solid", start_color=DISP, end_color=DISP), font=Font(color="FFFFFF", bold=True)))

# Cor da frente coluna E
ws.conditional_formatting.add(f"E2:E{N}", FormulaRule(formula=[f'EXACT($E2,"Doctum")'], fill=PatternFill(fill_type="solid", start_color=DOCTUM, end_color=DOCTUM), font=Font(color="FFFFFF", bold=True)))
ws.conditional_formatting.add(f"E2:E{N}", FormulaRule(formula=[f'EXACT($E2,"Tribo")'], fill=PatternFill(fill_type="solid", start_color=TRIBO, end_color=TRIBO), font=Font(color="FFFFFF", bold=True)))
ws.conditional_formatting.add(f"E2:E{N}", FormulaRule(formula=[f'EXACT($E2,"Dra. Bruna")'], fill=PatternFill(fill_type="solid", start_color=BRUNA, end_color=BRUNA), font=Font(color="FFFFFF", bold=True)))
ws.conditional_formatting.add(f"E2:E{N}", FormulaRule(formula=[f'EXACT($E2,"CuidaJA")'], fill=PatternFill(fill_type="solid", start_color=CUIDAJA, end_color=CUIDAJA), font=Font(color="FFFFFF", bold=True)))

# Freeze
ws.freeze_panes = "A2"

# ============================================
# ABA 4: ESPELHO
# ============================================
ws = wb.create_sheet("Espelho")

# Título
ws.merge_cells("A1:G1")
title = ws.cell(row=1, column=1, value="📊 O Espelho — Matriz Frente × Categoria (Realizado)")
title.font = Font(size=16, bold=True, color=HEADER_FG)
title.fill = header_fill()
title.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 32

# Matriz Realizado
ws.cell(row=3, column=1, value="").fill = soft_fill()
for i, (nome, icon, cor, orc) in enumerate(FRENTES):
    col_letter = get_column_letter(2 + i)
    c = ws.cell(row=3, column=2+i, value=f"{icon} {nome}")
    c.font = Font(bold=True, color=HEADER_FG)
    c.fill = PatternFill(fill_type="solid", start_color=cor, end_color=cor)
    c.alignment = Alignment(horizontal="center")
    c.border = border_all

total_col = 2 + len(FRENTES)
c = ws.cell(row=3, column=total_col, value="Total")
c.font = header_font()
c.fill = header_fill()
c.alignment = Alignment(horizontal="center")
c.border = border_all

# Linhas Bússola
triade_rows = [
    ("🎯 Importante", "Importante", IMP, 4),
    ("🔥 Urgente", "Urgente", URG, 5),
    ("💨 Disperso", "Disperso", DISP, 6),
]

for label, val, color, row in triade_rows:
    c = ws.cell(row=row, column=1, value=label)
    c.font = Font(bold=True, color=HEADER_FG)
    c.fill = PatternFill(fill_type="solid", start_color=color, end_color=color)
    c.alignment = Alignment(horizontal="center")
    c.border = border_all
    for i, (nome, icon, cor, orc) in enumerate(FRENTES):
        # SUMIFS(Blocos!D:D, Blocos!E:E, "Doctum", Blocos!H:H, "Importante")
        cell = ws.cell(row=row, column=2+i,
                       value=f'=SUMIFS(Blocos!D:D,Blocos!E:E,"{nome}",Blocos!H:H,"{val}")')
        cell.alignment = Alignment(horizontal="center")
        cell.number_format = '0.0" h"'
        cell.border = border_all
    # Total da linha
    rng = f'B{row}:{get_column_letter(total_col-1)}{row}'
    c = ws.cell(row=row, column=total_col, value=f'=SUM({rng})')
    c.alignment = Alignment(horizontal="center")
    c.number_format = '0.0" h"'
    c.font = Font(bold=True)
    c.fill = soft_fill()
    c.border = border_all

# Total geral
c = ws.cell(row=7, column=1, value="Total")
c.font = header_font()
c.fill = header_fill()
c.alignment = Alignment(horizontal="center")
c.border = border_all
for i, (nome, icon, cor, orc) in enumerate(FRENTES):
    col = 2 + i
    rng = f'{get_column_letter(col)}4:{get_column_letter(col)}6'
    cell = ws.cell(row=7, column=col, value=f'=SUM({rng})')
    cell.alignment = Alignment(horizontal="center")
    cell.number_format = '0.0" h"'
    cell.font = Font(bold=True)
    cell.fill = soft_fill()
    cell.border = border_all
c = ws.cell(row=7, column=total_col, value=f'=SUM(B7:{get_column_letter(total_col-1)}7)')
c.font = Font(bold=True, size=14)
c.fill = soft_fill()
c.alignment = Alignment(horizontal="center")
c.number_format = '0.0" h"'
c.border = border_all

# Larguras
ws.column_dimensions["A"].width = 22
for i in range(len(FRENTES)):
    ws.column_dimensions[get_column_letter(2+i)].width = 16
ws.column_dimensions[get_column_letter(total_col)].width = 14

# === Comparativo planejado vs realizado (linha 10+) ===
ws.cell(row=9, column=1, value="Planejado vs Realizado (% da semana)").font = Font(bold=True, size=13, color=ACCENT)

headers_comp = ["Categoria", "Planejado %", "Realizado %", "Variação", "Avaliação"]
for col, h in enumerate(headers_comp, start=1):
    c = ws.cell(row=10, column=col, value=h)
    c.font = header_font()
    c.fill = header_fill()
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = border_all
ws.row_dimensions[10].height = 24

# Fórmulas para cada categoria
comp_rows = [
    ("🎯 Importante", "Importante", IMP, 11),
    ("🔥 Urgente", "Urgente", URG, 12),
    ("💨 Disperso", "Disperso", DISP, 13),
]

for label, val, color, row in comp_rows:
    c = ws.cell(row=row, column=1, value=label)
    c.font = Font(bold=True, color=HEADER_FG)
    c.fill = PatternFill(fill_type="solid", start_color=color, end_color=color)
    c.alignment = Alignment(horizontal="center")
    c.border = border_all
    # Planejado %
    pl = f'=IFERROR(SUMIF(Blocos!G:G,"{val}",Blocos!D:D)/SUM(Blocos!D:D),0)'
    cp = ws.cell(row=row, column=2, value=pl)
    cp.number_format = "0%"
    cp.alignment = Alignment(horizontal="center")
    cp.border = border_all
    # Realizado %
    rl = f'=IFERROR(SUMIF(Blocos!H:H,"{val}",Blocos!D:D)/SUM(Blocos!D:D),0)'
    cr = ws.cell(row=row, column=3, value=rl)
    cr.number_format = "0%"
    cr.alignment = Alignment(horizontal="center")
    cr.border = border_all
    # Variação
    cv = ws.cell(row=row, column=4, value=f'=C{row}-B{row}')
    cv.number_format = "+0%;-0%;0%"
    cv.alignment = Alignment(horizontal="center")
    cv.border = border_all
    cv.font = Font(bold=True)
    # Avaliação
    if val == "Importante":
        av = f'=IF(D{row}>=0.05,"⬆ Bom sinal",IF(D{row}<=-0.05,"⬇ Atenção","✓ No plano"))'
    else:
        av = f'=IF(D{row}<=-0.05,"⬇ Bom sinal",IF(D{row}>=0.05,"⬆ Atenção","✓ No plano"))'
    ca = ws.cell(row=row, column=5, value=av)
    ca.alignment = Alignment(horizontal="center")
    ca.border = border_all
    ca.font = Font(bold=True)

ws.column_dimensions["B"].width = 16
ws.column_dimensions["C"].width = 16
ws.column_dimensions["D"].width = 14
ws.column_dimensions["E"].width = 18

# === Resumo por frente: orçamento vs realizado ===
ws.cell(row=15, column=1, value="Orçamento de horas por frente").font = Font(bold=True, size=13, color=ACCENT)

headers_orc = ["Frente", "Orçamento", "Realizado", "Variação", "Status"]
for col, h in enumerate(headers_orc, start=1):
    c = ws.cell(row=16, column=col, value=h)
    c.font = header_font()
    c.fill = header_fill()
    c.alignment = Alignment(horizontal="center")
    c.border = border_all
ws.row_dimensions[16].height = 24

for i, (nome, icon, cor, orc) in enumerate(FRENTES):
    row = 17 + i
    c = ws.cell(row=row, column=1, value=f"{icon} {nome}")
    c.font = Font(bold=True, color=HEADER_FG)
    c.fill = PatternFill(fill_type="solid", start_color=cor, end_color=cor)
    c.alignment = Alignment(horizontal="center")
    c.border = border_all
    # Orçamento (vlookup da aba Frentes)
    cc = ws.cell(row=row, column=2, value=f'=VLOOKUP(SUBSTITUTE(SUBSTITUTE(A{row},"{icon} ",""),"",""),Frentes!A:C,3,FALSE)')
    cc.alignment = Alignment(horizontal="center")
    cc.number_format = '0.0" h"'
    cc.border = border_all
    # Realizado
    cd = ws.cell(row=row, column=3, value=f'=SUMIF(Blocos!E:E,"{nome}",Blocos!D:D)')
    cd.alignment = Alignment(horizontal="center")
    cd.number_format = '0.0" h"'
    cd.border = border_all
    # Variação
    ce = ws.cell(row=row, column=4, value=f'=C{row}-B{row}')
    ce.alignment = Alignment(horizontal="center")
    ce.number_format = '+0.0" h";-0.0" h";0" h"'
    ce.border = border_all
    ce.font = Font(bold=True)
    # Status
    cf = ws.cell(row=row, column=5, value=f'=IF(ABS(D{row})<1,"✓ no plano",IF(D{row}>0,"⬆ estourou","⬇ abaixo"))')
    cf.alignment = Alignment(horizontal="center")
    cf.border = border_all
    cf.font = Font(bold=True)

# Freeze
ws.freeze_panes = "A4"

# ============================================
# ABA 5: INSIGHTS
# ============================================
ws = wb.create_sheet("Insights")

# Título
ws.merge_cells("A1:B1")
c = ws.cell(row=1, column=1, value="💬 Insights do Coach")
c.font = Font(size=16, bold=True, color=HEADER_FG)
c.fill = header_fill()
c.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 32

ws.cell(row=2, column=1, value="Padrões detectados automaticamente · Recalcula ao mudar Blocos").font = Font(italic=True, color="64748B")
ws.row_dimensions[2].height = 22

ws.column_dimensions["A"].width = 30
ws.column_dimensions["B"].width = 80

# Headers
c = ws.cell(row=4, column=1, value="Sinal")
c.font = header_font()
c.fill = header_fill()
c.alignment = Alignment(horizontal="center")
c.border = border_all
c = ws.cell(row=4, column=2, value="Insight")
c.font = header_font()
c.fill = header_fill()
c.alignment = Alignment(horizontal="center")
c.border = border_all
ws.row_dimensions[4].height = 24

# Fórmulas para cada insight
# Variáveis úteis:
TOTAL = "SUM(Blocos!D:D)"
TOT_IMP = 'SUMIF(Blocos!H:H,"Importante",Blocos!D:D)'
TOT_URG = 'SUMIF(Blocos!H:H,"Urgente",Blocos!D:D)'
TOT_DISP = 'SUMIF(Blocos!H:H,"Disperso",Blocos!D:D)'

PCT_IMP = f'({TOT_IMP}/{TOTAL})'
PCT_URG = f'({TOT_URG}/{TOTAL})'
PCT_DISP = f'({TOT_DISP}/{TOTAL})'

insights_rows = [
    ("Importante - viveu em estratégia",
     f'=IF({PCT_IMP}>=0.5,"✅ Viveu em Importante","")',
     f'=IF({PCT_IMP}>=0.5,TEXT({PCT_IMP},"0%")&" da semana foi Importante. Esse é o sinal de uma semana onde estratégia venceu reatividade. Mantenha o ritmo.","")'),
    ("Importante - pouco protegido",
     f'=IF({PCT_IMP}<0.3,"⚠️ Pouco Importante","")',
     f'=IF({PCT_IMP}<0.3,"Apenas "&TEXT({PCT_IMP},"0%")&" da semana foi Importante. Sua estratégia está perdendo pra urgência alheia. Vale revisar o que está protegendo deep work.","")'),
    ("Urgente - semana de bombeiro",
     f'=IF({PCT_URG}>=0.5,"🔥 Bombeiro","")',
     f'=IF({PCT_URG}>=0.5,TEXT({PCT_URG},"0%")&" Urgente. Você está reagindo, não decidindo. Olhe quais frentes mais consumiram Urgente — provavelmente dá pra delegar ou redesenhar pipeline.","")'),
    ("Disperso alto",
     f'=IF({PCT_DISP}>=0.2,"💨 Atenção","")',
     f'=IF({PCT_DISP}>=0.2,TEXT({PCT_DISP},"0%")&" Disperso = "&TEXT({TOT_DISP},"0.0")&"h de coisa que pareceu trabalho mas não gerou resultado. Vale cortar reuniões/mensagens sem propósito.","")'),
    ("Disperso baixo",
     f'=IF({PCT_DISP}<0.1,"🎯 Foco","")',
     f'=IF({PCT_DISP}<0.1,"Apenas "&TEXT({PCT_DISP},"0%")&" Disperso — você tá protegendo bem o tempo de coisa que não rende.","")'),
]

current_row = 5
for label, sinal_f, insight_f in insights_rows:
    cs = ws.cell(row=current_row, column=1, value=sinal_f)
    cs.font = Font(bold=True, size=12)
    cs.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cs.border = border_all
    ci = ws.cell(row=current_row, column=2, value=insight_f)
    ci.alignment = Alignment(vertical="center", wrap_text=True)
    ci.border = border_all
    ws.row_dimensions[current_row].height = 48
    current_row += 1

# Por frente: insight se Urgente > 60% naquela frente
ws.cell(row=current_row, column=1, value="Por frente").font = Font(bold=True, italic=True, color=ACCENT)
current_row += 1

for nome, icon, cor, orc in FRENTES:
    sf_total = f'SUMIF(Blocos!E:E,"{nome}",Blocos!D:D)'
    sf_urg = f'SUMIFS(Blocos!D:D,Blocos!E:E,"{nome}",Blocos!H:H,"Urgente")'
    sf_imp = f'SUMIFS(Blocos!D:D,Blocos!E:E,"{nome}",Blocos!H:H,"Importante")'

    sinal_f = f'=IF({sf_total}=0,"",IF(({sf_urg}/{sf_total})>=0.6,"⚠️ Bombeiro",IF(({sf_imp}/{sf_total})>=0.7,"🎯 Protegida","")))'
    insight_f = (f'=IF({sf_total}=0,"Sem horas em {icon} {nome} esta semana.",'
                 f'IF(({sf_urg}/{sf_total})>=0.6,"{icon} {nome}: "&TEXT({sf_urg}/{sf_total},"0%")&" Urgente. Vale pensar em delegar, redesenhar pipeline de demanda ou aceitar conscientemente que essa é sua frente reativa.",'
                 f'IF(({sf_imp}/{sf_total})>=0.7,"{icon} {nome}: "&TEXT({sf_imp}/{sf_total},"0%")&" Importante. Você está investindo bem nessa frente — onde provavelmente está seu maior ROI estratégico.",'
                 f'"{icon} {nome}: equilibrada ("&TEXT({sf_imp}/{sf_total},"0%")&" Imp / "&TEXT({sf_urg}/{sf_total},"0%")&" Urg).")))')

    cs = ws.cell(row=current_row, column=1, value=sinal_f)
    cs.font = Font(bold=True, size=12)
    cs.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cs.border = border_all
    ci = ws.cell(row=current_row, column=2, value=insight_f)
    ci.alignment = Alignment(vertical="center", wrap_text=True)
    ci.border = border_all
    ws.row_dimensions[current_row].height = 48
    current_row += 1

# Rodapé com dica
current_row += 1
c = ws.cell(row=current_row, column=1, value="💡 Coach gentil V1")
c.font = Font(bold=True, italic=True, color=ACCENT)
ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
c = ws.cell(row=current_row + 1, column=1, value="Os insights são gerados por fórmulas IF baseadas em regras simples. Mude qualquer coisa na aba Blocos e os textos atualizam automaticamente.")
c.font = Font(italic=True, color="64748B", size=10)
c.alignment = Alignment(wrap_text=True)
ws.merge_cells(start_row=current_row + 1, start_column=1, end_row=current_row + 1, end_column=2)
ws.row_dimensions[current_row + 1].height = 30

# ============================================
# Reordenar abas
# ============================================
ordem = ["Sobre", "Frentes", "Blocos", "Espelho", "Insights"]
wb._sheets = [wb[name] for name in ordem]

# Aba ativa = Blocos
wb.active = wb.sheetnames.index("Blocos")

wb.save(OUT)
print(f"OK gerado: {OUT}")
